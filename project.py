import numpy as np
import pandas as pd
from openpyxl import load_workbook
from random import choice
import matplotlib.pyplot as plt


POPULATION_SIZE = 500
CROSS_RATE = 0.6
MUTATION_RATE = 0.06
GENERATIONS = 200

# N = 10  # number of students
CHROMOSOME_LENGTH = 15


# excel_name = 'input.xlsx'
# d = pd.DataFrame(pd.read_excel(excel_name, usecols='A,F,G,H,AA', names=['ID', 'SEN', 'PRE', 'DoB', 'Hours'], skiprows=[0, 1, 2, 3]))
# d = d.dropna(subset=['ID'])
# d.fillna(0, inplace=True)
# d2 = d.sort_values(by="DoB", ascending=True, inplace=False)

replace_list = np.array([0, 2, 3, 5, 6, 8, 9, 11, 12, 14])


class GA:
    def __init__(self, N, chromlen, cross_rate, mutation_rate, pop_size, constrains, col_constrain):
        self.students = N
        self.chromlen = chromlen
        self.cross_rate = cross_rate
        self.mutate_rate = mutation_rate
        self.pop_size = pop_size
        self.constrains = constrains
        self.col_constrain = col_constrain

        self.pop = np.random.randint(2, size=(pop_size, N, chromlen))

    def fitness(self):
        result = np.zeros(self.pop_size)
        for j in range(self.pop_size):
            table = self.pop[j]
            for i in range(N):
                line = table[i]
                totalhours = 0
                perference = []
                timeslots = np.split(line, 5)
                for num in range(5):
                    ary = timeslots[num]
                    hours = self.cal_timeslots(ary)
                    if hours != -1:
                        totalhours += hours
                    else:
                        totalhours += 1.25
                        if result[j] - 10 > 0:
                            result[j] -= 10
                        else:
                            result[j] = 0
                    perference.append(10*(num+1)+self.cal_perference(ary))
                if self.constrains[i][2] != 0:
                    if self.constrains[i][2] in perference:
                        result[j] += self.constrains[i][5]
                else:
                    result[j] += 1
                if totalhours == self.constrains[i][4]:
                    result[j] += 20
                elif totalhours < self.constrains[i][4]:
                    result[j] += 1
                else:
                    if result[j] - 5 > 0:
                        result[j] -= 5
                    else:
                        result[j] = 0
            p = 0
            for col in table.T:
                if np.sum(col) + col_constrain[p] > 26:
                    result[j] = 0
                p = p + 1
        return result

    # convert an array like [1,0,1] to a string like '101'
    # then convert this string to a binary number 101
    def convertArrary(self, ary):
        string = ''
        for num in ary:
            string += str(num)
        bin2int = int(string, 2)
        return bin2int

    # calculate the binary number like 101
    def cal_perference(self, ary):
        bin2int = self.convertArrary(ary)
        if bin2int == 1 or bin2int == 3:
            result = 2
        elif bin2int == 6 or bin2int == 4:
            result = 1
        else:
            result = 0
        return result

    # calculate the total hours of each child by the binary number like 101
    def cal_timeslots(self, ary):
        bin2int = self.convertArrary(ary)
        if bin2int == 0:
            result = 0
        elif bin2int == 1 or bin2int == 4:
            result = 2.5
        elif bin2int == 3 or bin2int == 6:
            result = 3.75
        elif bin2int == 5:
            result = 5
        elif bin2int == 7:
            result = 6.25
        else:
            result = -1
        return result

    # https://github.com/MorvanZhou/Evolutionary-Algorithm/blob/master/tutorial-contents/Genetic%20Algorithm/Match%20Phrase.py
    # provided idea about the selection function for me
    def select(self):
        fitness = self.fitness() + 0.0000001
        # print(fitness)
        idx = np.random.choice(np.arange(self.pop_size), size=self.pop_size, replace=True, p=fitness/fitness.sum())
        return self.pop[idx]

    def crossover(self, parent, pop):
        if np.random.rand() < self.cross_rate:
            random_individual = np.random.randint(0, self.pop_size, size=1)
            cross_points = np.random.randint(0, self.students, round(self.students / 2))
            parent[cross_points] = pop[random_individual, cross_points]
        return parent

    def mutate(self, child):
        for point in replace_list:
            if np.random.rand() < self.mutate_rate:
                child.T[point] = abs(child.T[point] - 1)
        return child

    def evolve(self):
        pop = self.select()
        pop_copy = pop.copy()
        for parent in pop:
            child = self.crossover(parent, pop_copy)
            child = self.mutate(child)
            parent[:] = child
        self.pop = pop


def generate_hours(n, r):
    result = []
    candidates = [2.5, -2.5, 1.25, -1.25]
    for i in range(n):
        if np.random.rand() < r:
            result.append(15+choice(candidates))
        else:
            result.append(15)
    return result

if __name__ == '__main__':
    # book = load_workbook('input.xlsx')
    # d = pd.DataFrame(generate_hours(33, 0.6))
    # writer = pd.ExcelWriter('input.xlsx', engine='openpyxl')
    # writer.book = book
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # d.to_excel(writer, startrow=26, startcol=26, header=False, index=False)
    # writer.save()



    # First part
    book = load_workbook('input.xlsx')
    excel_name = 'input.xlsx'
    d = pd.DataFrame(pd.read_excel(excel_name, usecols='A,F,G,H,AA', names=['ID', 'SEN', 'PRE', 'DoB', 'Hours'],
                                   skiprows=[0, 1, 2, 3]))
    d = d.dropna(subset=['ID'])
    d.fillna(0, inplace=True)

    d3 = d.loc[21:39]
    d3 = d3.sort_values(by="DoB", ascending=True, inplace=False)
    d5 = np.arange(0.0, 10.0, 10 / 19).round(2)[::-1]
    x = d3['SEN'] * 10
    d3['Weight'] = d5 + x
    N = d3.shape[0]
    d6 = d3.sort_values(by="ID", ascending=True, inplace=False)
    constrains = np.array(d6)
    schdule = np.array(pd.read_excel('input.xlsx', usecols='K:Y', skiprows=[0, 1, 2, 3]))
    schdule[np.isnan(schdule)] = 0
    col_constrain = np.sum(schdule, axis=0)

    ga = GA(N=N, chromlen=CHROMOSOME_LENGTH, cross_rate=CROSS_RATE,
            mutation_rate=MUTATION_RATE, pop_size=POPULATION_SIZE,
            constrains=constrains, col_constrain=col_constrain)
    autumn = []
    for generation in range(GENERATIONS):
        ga.evolve()
        fitness = ga.fitness()
        autumn.append(max(fitness))


    fitness = ga.fitness()
    maxfitness = max(fitness)
    idxn = np.argmax(fitness)
    best = ga.pop[idxn]
    data = pd.DataFrame(best)
    writer = pd.ExcelWriter('input.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, startrow=26, startcol=10, header=False, index=False)
    writer.save()

    #Second Part
    book = load_workbook('input.xlsx')

    d3 = d.loc[41:45]
    d3 = d3.sort_values(by="DoB", ascending=True, inplace=False)
    d5 = np.arange(0.0, 10.0, 10 / 5).round(2)[::-1]
    x = d3['SEN'] * 10
    d3['Weight'] = d5 + x
    N = d3.shape[0]
    d6 = d3.sort_values(by="ID", ascending=True, inplace=False)
    constrains = np.array(d6)
    schdule = np.array(pd.read_excel('input.xlsx', usecols='K:Y', skiprows=[0, 1, 2, 3]))
    schdule[np.isnan(schdule)] = 0
    col_constrain = np.sum(schdule, axis=0)
    ga = GA(N=N, chromlen=CHROMOSOME_LENGTH, cross_rate=CROSS_RATE,
            mutation_rate=MUTATION_RATE, pop_size=POPULATION_SIZE,
            constrains=constrains, col_constrain=col_constrain)
    spring = []
    for generation in range(GENERATIONS):
        ga.evolve()
        fitness = ga.fitness()
        spring.append(max(fitness))


    fitness = ga.fitness()
    maxfitness = max(fitness)
    idxn = np.argmax(fitness)
    best = ga.pop[idxn]
    data = pd.DataFrame(best)
    writer = pd.ExcelWriter('input.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, startrow=46, startcol=10, header=False, index=False)
    writer.save()

    #Third Part
    book = load_workbook('input.xlsx')

    d3 = d.loc[47:54]
    d3 = d3.sort_values(by="DoB", ascending=True, inplace=False)
    d5 = np.arange(0.0, 10.0, 10 / 8).round(2)[::-1]
    x = d3['SEN'] * 10
    d3['Weight'] = d5 + x
    N = d3.shape[0]
    d6 = d3.sort_values(by="ID", ascending=True, inplace=False)
    constrains = np.array(d6)
    schdule = np.array(pd.read_excel('input.xlsx', usecols='K:Y', skiprows=[0, 1, 2, 3]))
    schdule[np.isnan(schdule)] = 0
    col_constrain = np.sum(schdule, axis=0)
    ga = GA(N=N, chromlen=CHROMOSOME_LENGTH, cross_rate=CROSS_RATE,
            mutation_rate=MUTATION_RATE, pop_size=POPULATION_SIZE,
            constrains=constrains, col_constrain=col_constrain)
    summer = []
    for generation in range(GENERATIONS):
        ga.evolve()
        fitness = ga.fitness()
        summer.append(max(fitness))

    fitness = ga.fitness()
    maxfitness = max(fitness)
    idxn = np.argmax(fitness)
    best = ga.pop[idxn]
    data = pd.DataFrame(best)
    writer = pd.ExcelWriter('input.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, startrow=52, startcol=10, header=False, index=False)
    writer.save()

    # book = load_workbook('input.xlsx')
    # d = pd.DataFrame(pd.read_excel(excel_name, usecols='K:Y', skiprows=range(1, 26)))
    # x = np.array(d)
    # d.iloc[:, :] = np.nan
    # writer = pd.ExcelWriter('input.xlsx', engine='openpyxl')
    # writer.book = book
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # d.to_excel(writer, startrow=26, startcol=10, header=False, index=False)
    # writer.save()

    x = np.linspace(0, 200, 200)
    plt.plot(x, np.array(autumn), label='autumn')
    plt.plot(x, np.array(spring), label='spring')
    plt.plot(x, np.array(summer), label='summer')

    plt.xlabel('Generation')
    plt.ylabel('Fitness Score')


    plt.legend()

    plt.show()
